#openpyxl库只能操作xlsx格式的表格
import shutil #复制文件
from openpyxl  import load_workbook#原有表格操作
from openpyxl.worksheet.copier import WorksheetCopy


class excle_xlsx():
    def __init__(self):
        self.old_file_name = 'modle.xlsx'
        self.modle_sheet_name = 'modle'
        self.new_file_name = '店小蜜数据.xlsx'

    def copy_workbook1(self):
        shutil.copyfile(self.old_file_name,self.new_file_name)

    def copy_sheet(self,new_sheet_name):
        '''复制一个sheet,执行此方法之后获得类变量self.wb,self.ws即可往表格内写数据,执行完成记得保存关闭'''
        self.wb = load_workbook(self.new_file_name)
        ws_modle = self.wb[self.modle_sheet_name]
        ws = self.wb.create_sheet(title=new_sheet_name,index=None)
        cp = WorksheetCopy(source_worksheet=ws_modle,target_workseet=ws)

    def save_close(self):
        self.wb.save(self.new_file_name)
        self.wb.close()

    def remove_sheet(self):
        self.wb.remove_sheet('self.modle_sheet_name ')

#wb = load_workbook(self.new_file_name)
#wb.copy_worksheet(wb.worksheets[0])



# file_name='DXM数据表.xlsx'
#print(file_name)
#shutil.copyfile(file_name, 'DXM数据表1.xlsx')
# wb=load_workbook(file_name)
# source_ws = wb['Sheet2']
# ws_1=wb.create_sheet(title='001',index=0)
# cp = WorksheetCopy(source_worksheet=source_ws,target_worksheet=ws_1)
# cp.copy_worksheet()
# wb.save('DXM数据表2.xlsx')
# copy_sheet1=wb.copy_worksheet(wb['Sheet2'],0)
# copy_sheet2=wb.copy_worksheet(wb['Sheet2'],0)
# copy_sheet3=wb.copy_worksheet(wb['Sheet2'],0)
# copy_sheet1.title='wei1'
# copy_sheet2.title='wei2'
# copy_sheet3.title='wei666'
# wb.save('DXM数据表1.xlsx')
# print(666)
# ws=wb['自然之名-店小蜜数据'] #得到sheet1
# ws['a33']='aaaaaaa'
# ws['b33']='aaaaaaa'
# ws['c33']='aaaaaaa'
# wb.save(file_name)
# wb.close()

#    def write_xlsx(self):
#        file_name = 'DXM数据表 展示.xlsx'
#        wb = load_workbook(file_name)
#        ws = wb['自然之名-店小蜜数据']
#        ws['a33'] = 'aaaa'
#        ws['b33'] = 'aaaa'
#        ws['c33'] = 'aaaa'
#        ws['d33'] = 'aaaa'
#        wb.save(file_name)
#        wb.close()



select * from tab1    where   between value1 and value2
