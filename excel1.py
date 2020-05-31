from openpyxl import load_workbook
import sqlit3_wei
import re

#本文件将表格数据写入数据库
filename = 'DXM数据表1.xlsx'

wb = load_workbook(filename)
print(7777888)
ws = wb['自然之名旗舰店']
for i in range(6,384):
    li=[j.value for j in ws[i]]

li=[[j.value for j in ws[i][:3]] for i in range(3,382)]
li1=[i for i in li if re.findall('(\d{4})-(\d{2})-(\d{2})',str(i[0]))]
li2=[]
for i in li1:
    rq=re.findall('(\d{4})-(\d{2})-(\d{2})',str(i[0]))[0]
    rq=''.join(rq)
    li2 +=[[int(rq),i[2]]]
dianpu='自然之名旗舰店'
sqlit3_1 = sqlit3_wei.weiSqlite3_dxm('data.db')
sqlit3_1.add_table1(dianpu)
for i in li2:
    sql1=f"update {dianpu}_排名 set 行业排名 = {i[1]} where 日期={i[0]}"
    try:
        sqlit3_1.s_add_del_dpd(sql1)
    except:print(i)



# filename = 'DXM数据表 - 副本1.xlsx'
#
# wb = load_workbook(filename)
# print(7777888)
# ws = wb['七匹狼箱包旗舰店']
# for i in range(3,382):
#     li=[j.value for j in ws[i]]
#
# li=[[j.value for j in ws[i]] for i in range(3,382)]
# li1=[i for i in li if re.findall('(\d{4})-(\d{2})-(\d{2})',str(i[0]))]
# li2=[]
# li3=[]
# for i in li1:
#     rq=re.findall('(\d{4})-(\d{2})-(\d{2})',str(i[0]))[0]
#     rq=''.join(rq)
#     li2 +=[[int(rq),i[1],i[2],i[4],i[5],i[10],i[11],i[7],i[9],i[8]]]
#     li3 += [[int(rq),i[6],i[13],i[12],i[3]]]
# print(111222333)
# dianpu='七匹狼箱包旗舰店'
# sqlit3_1 = sqlit3_wei.weiSqlite3_dxm('data.db')
# sqlit3_1.add_table1(dianpu)
# for i in li2:
#     sql1=f"Insert Into {dianpu}_排名 values(?,?,?,?,?,?,?,?,?,?)"
#     #sql1=f"Insert Into {dianpu}_排名 (行业排名) Values({value1})"
#     try:
#         sqlit3_1.s_add_del_dpd(sql1,i)
#     except:print(i)

# for i1 in li3:
#     sql2=f"Insert Into {dianpu}_转化 values(?,?,?,?,?)"
#     try:
#         sqlit3_1.s_add_del_dpd(sql2,i1)
#     except:print(i1)




print(66666666)
# e12=sqlit3_1.s_sel('select * from 自然之名旗舰店_转化')
# e13=sqlit3_1.s_sel('select * from 自然之名旗舰店_排名')
sqlit3_1.conn.close()
