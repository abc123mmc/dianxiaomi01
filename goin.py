from selenium import webdriver
import pyautogui
import time
import os
import pyperclip
import re
import requests
import json
import yaml
import params
import sqlit3_wei
import shutil #复制文件
from openpyxl  import load_workbook#原有表格操作
from openpyxl.worksheet.copier import WorksheetCopy

class get_data_dxm():
    def __init__(self,user_name,pass_words):
        self.user_name = user_name
        time.sleep(3)
        option = webdriver.ChromeOptions()
        option.add_argument(r'--user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data') 
        self.driver=webdriver.Chrome(options=option)
        self.driver.maximize_window()
        self.driver.get('https://login.taobao.com')
        time.sleep(3)
        pyautogui.press('tab')
        time.sleep(1)
        pyautogui.press('tab')
        pyautogui.press('enter')
        pyperclip.copy(user_name)
        time.sleep(1)
        pyautogui.hotkey('ctrl','v')
        time.sleep(1)
        pyautogui.press('tab')
        pyautogui.typewrite(pass_words)
        time.sleep(1)
        pyautogui.press('enter')
        while self.driver.current_url != 'https://mai.taobao.com/seller_admin.htm':
            time.sleep(3)
        self.driver.get('https://dianxiaomi.taobao.com')
        time.sleep(2)
        self.driver.find_element_by_xpath('//a[@data-title="全自动数据"]').click()

    def get_cook(self):
        cookies1=self.driver.get_cookies()
        cookies1=[f'{i["name"]}={i["value"]}' for i in cookies1]
        return ';'.join(cookies1)
        
    def get_data_req(self,param1):
        '''传入请求参数,返回请求到的数据
        : param param1 : 请求数据类型需要的数据,从params文件获取,字典类型
        : return data : 请求到的数据,第一个元素为名称
        : rtype :列表
        '''
        cooks1=self.get_cook() #获取cook
        headers = params.headers
        headers['cookie'] = cooks1 #将cook写入到headers中
        ddict = params.ddict
        tb_token = re.findall('tb_token_=([^;]+)',cooks1)[0] #从cooks1中提取token
        ddict['_tb_token_'] = tb_token #设置ddict(请求参数)
        ddict['_dxm_token_'] = tb_token
        ddict['param'] = json.dumps(param1)  #json.loads(str)

        res = requests.post(params.url,headers=headers,data=ddict)
        e = res.json() #获取返回数据并做处理并返回
        e1 = json.loads(e['data']['xinsight'])
        e11 = e1['data']['cellset']
        data=[[j['raw'] for j in i] for i in e11]

        return data

    def main(self,start_time,end_time):
        param1 = params.param_paiming
        param1['startTime'] = start_time
        param1['endTime'] = end_time
        self.data1 = self.get_data_req(param1)
        self.sqlit3_1 = sqlit3_wei.weiSqlite3_dxm('data.db')
        self.sqlit3_1.add_table1(self.user_name)
        for i in self.data1[1:]:
            li=[int(i[0])] + [j for j in i[1:]]
            try:self.sqlit3_1.s_add_del_dpd(f"Insert Into {self.sqlit3_1.paiming} values(?,?,?,?,?,?,?,?,?,?,?)",li)
            except:
                print(li[0],85)

        param2 = params.param_zhuanhua
        param2['startTime'] = start_time
        param2['endTime'] =end_time
        self.data2 = self.get_data_req(param2)
        for i in self.data2[1:]:
            li=[int(i[0])] + [j for j in i[1:-2]]+['%.4f' %(i[-1]/i[-2])] #小数保留4位
            try:self.sqlit3_1.s_add_del_dpd(f"Insert Into {self.sqlit3_1.zhuanhua} values(?,?,?,?,?)",li)
            except:
                print(li[0],95)

    def previous_months(self,num):
        t=time.localtime()
        y,m=t[:2]
        li=[]
        for i in range(num):
            li += [[y,m]]
            if m == 1:
                m = 12
                y -=1
            else:
                m -=1
        return li[::-1]



if __name__ == '__main__':
    old_file_name = 'modle.xlsx'
    modle_sheet_name = 'modle'
    new_file_name = '店小蜜数据.xlsx'
    with open(r'配置文件.yml','r',encoding="gbk") as f:
        data_zh = yaml.load(f,Loader=yaml.FullLoader)
    shutil.copyfile(old_file_name,new_file_name)
    for i in data_zh:
        print(i)
        wb = load_workbook(new_file_name)
        ws_modle = wb[modle_sheet_name]
        ws = wb.create_sheet(title=i.split(':')[0],index=None)
        cp=WorksheetCopy(source_worksheet=ws_modle,target_worksheet=ws)
        cp.copy_worksheet()
        wb.save(new_file_name)
        wb.close()
    wb = load_workbook(new_file_name)
    wb.remove(wb['modle'])
    wb.save(new_file_name)
    wb.close()
    for i in data_zh:
        os.system('taskkill /im chromedriver.exe /F')
        os.system('taskkill /im chrome.exe /F')
        time.sleep(2)
        user_name,pass_words = i,data_zh[i]
        rq1 = 0
        try:
            sq=sqlit3_wei.weiSqlite3('data.db')
            rq_zj=sq.s_sel(f'select max(cast(日期 as int)) from  {i.split(":")[0]}_排名')[0][0]
            print(rq_zj)
            rq1=str(rq_zj)
            rq1= time.mktime((int(rq1[:4]),int(rq1[4:6]),int(rq1[6:8]),0,0,0,0,0,0))
            print(rq1)
        except:pass
        t = time.time()
        if rq1>t-3600*24*2:
            star_time='无需获取'
        elif 3600*24*30<rq1<t-3600*24*2:
            t1=time.localtime(rq1+3600*24)
            t2=time.localtime(t-3600*24)
            star_time=f'{t1[0]}-{"%02d"% t1[1]}-{"%02d"% t1[2]} 23:59:59'
            end_time = f'{t2[0]}-{"%02d"% t2[1]}-{"%02d"% t2[2]} 23:59:59'
        else:
            t1=time.localtime(t-3600*24*30)
            t2=time.localtime(t-3600*24)
            star_time=f'{t1[0]}-{"%02d"% t1[1]}-{"%02d"% t1[2]} 23:59:59'
            end_time = f'{t2[0]}-{"%02d"% t2[1]}-{"%02d"% t2[2]} 23:59:59'

        b=get_data_dxm(user_name,pass_words)
        if star_time != '无需获取':
            b.main(star_time,end_time)
        li=b.previous_months(12)
        row=[i2 for i2 in range(3,399,33)]
        wb = load_workbook(new_file_name)
        ws = wb[i.split(':')[0]]
        for i2,i1 in zip(li,row):
            sql1=f'''select * 
            from {i.split(':')[0]}_排名 pm,{i.split(':')[0]}_转化 zh 
            where pm.日期=zh.日期 and (pm.日期 between {i2[0]}{'%02d'% i2[1]}00 and {i2[0]}{'%02d'% i2[1]}32)'''
            sqlit3_2 = sqlit3_wei.weiSqlite3_dxm('data.db')
            row_v=sqlit3_2.s_sel(sql1)
            for i6 in row_v:
                ws[f'a{i1}']=i6[0]
                ws[f'e{i1}'] = i6[1]
                ws[f'f{i1}'] = i6[2]
                ws[f'j{i1}'] = i6[3]
                ws[f'k{i1}'] = i6[4]
                ws[f'u{i1}'] = i6[5]
                ws[f'v{i1}'] = i6[6]
                ws[f'p{i1}'] = i6[7]
                ws[f's{i1}'] = i6[8]
                ws[f'q{i1}'] = i6[9]
                ws[f'c{i1}'] = i6[10]

                ws[f'o{i1}'] = i6[12]
                ws[f'x{i1}'] = i6[13]
                ws[f'w{i1}'] = i6[14]
                ws[f'h{i1}'] = i6[15]
                i1 +=1
        wb.save(new_file_name)
        wb.close()
        b.driver.quit()



    # li=['日期','行业排名','转人工率','转人工率行业前三','实际转人工数量',
    #     '人工衔接后下单率','卡片点击率','卡片点击率行业前三','纯店小蜜询单转化',
    #     '小蜜参与询单转化','人工参与询单转化','转化率差值','转化率行业前三',
    #     '接待占比','使用情况行业前三','小蜜参与询单人数','接待人数','人工询单人数',
    #     '人工转化人数','小蜜转化人数','月实际转化率','增益差值']


    #e111=b.sqlit3_1.s_sel('select * from 自然之名旗舰店_排名')
    #e112=b.sqlit3_1.s_sel('select * from 自然之名旗舰店_转化')

